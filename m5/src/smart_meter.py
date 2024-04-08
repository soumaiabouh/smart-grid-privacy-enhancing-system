from openpyxl import load_workbook
from Crypto.Cipher import AES
from Crypto.Random import get_random_bytes
from Crypto.PublicKey import RSA
from Crypto.Cipher import PKCS1_OAEP

class SmartMeter:
    def __init__(self, pes_public_key, filename):
        self.pes_public_key = RSA.import_key(pes_public_key)
        self.aes_key = self._generate_key()
        self.encrypted_aes_key = self._encrypt_aes_key()
        self.id = self._generate_id()
        self.encrypted_data_list = []
        self._load_data(filename, 1, 60) # Initially loading 1 hour worth of data

    def _load_data(self, filename, start_row=1, end_row=None):
        try:
            workbook = load_workbook(filename=filename)
            sheet = workbook.active
            
            for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                if i < start_row or (end_row and i > end_row):  
                    continue  # Skip rows outside the specified range
                # Assuming your data is in the first two columns
                timestamp, reading = row[0], row[1]
                self._encrypt_data_and_store(timestamp, reading)  # Encrypt each reading with its timestamp
        except FileNotFoundError:
            print(f"The file {filename} was not found.")
        except Exception as e:
            print(f"An error occurred: {e}")

        
    def _generate_key(self):
        # Generate a random 256-bit key
        return get_random_bytes(32)
    
    def _generate_id(self):
        # generate id and return it
        return get_random_bytes(16)
        
    def _encrypt_data_and_store(self, timestamp, reading):
        cipher_aes = AES.new(self.aes_key, AES.MODE_GCM)  # Important to generate a new cipher every time for security reasons
        
        # Ensure reading is in a format suitable for encryption (i.e., bytes)
        if isinstance(reading, (int, float)):
            reading = str(reading).encode('utf-8')  # Convert numerical values to strings, then to bytes
        elif isinstance(reading, str):
            reading = reading.encode('utf-8')  # Encode string values directly
        else:
            raise ValueError("Unsupported data type for encryption")
        
        encrypted_data, tag = cipher_aes.encrypt_and_digest(reading)  # Encrypt the data
        nonce = cipher_aes.nonce  # Nonce is needed for decryption in GCM mode
        
        self.encrypted_data_list.append({
            'timestamp': timestamp,
            'encrypted_data': encrypted_data,
            'nonce': nonce,
            'tag': tag
        })
        
    def _encrypt_aes_key(self):
        # function that encrypts using RSA the smart meter AES key using a public key 
        cipher_rsa = PKCS1_OAEP.new(self.pes_public_key)
        encrypted_key = cipher_rsa.encrypt(self.aes_key)
        return encrypted_key

    def get_encrypted_data(self):
        return self.encrypted_data_list
            
    def get_id(self):
        return self.id
    
    def get_encrypted_aes_key(self):
        return self.encrypted_aes_key

    def generate_data(self, filename, timerange=60):
        # Determine the next set of rows to load based on the current data list length
        start_row = len(self.encrypted_data_list) + 1
        end_row = start_row + timerange - 1  # Adjust the end row based on the timerange
        
        # Load and encrypt the new range of data
        self._load_data(filename, start_row=start_row, end_row=end_row)
